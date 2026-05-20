import os
import time
import joblib
import warnings
import numpy as np
import pandas as pd

import matplotlib
matplotlib.use('Agg')

import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import seaborn as sns

from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import StratifiedKFold, cross_validate
from sklearn.metrics import (
    confusion_matrix,
    accuracy_score,
    f1_score,
    roc_auc_score
)

from sklearn.linear_model import LogisticRegression
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier
from sklearn.svm import SVC
from sklearn.neighbors import KNeighborsClassifier
from sklearn.naive_bayes import GaussianNB
from xgboost import XGBClassifier

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

warnings.filterwarnings('ignore')


# =========================================================
# PATHS
# =========================================================

BASE_DIR = r"C:\Users\emade\Downloads\Dementia_Detection"

SPLIT_DIR   = os.path.join(BASE_DIR, "data", "splits")
RESULTS_DIR = os.path.join(BASE_DIR, "results")
VISUALS_DIR = os.path.join(RESULTS_DIR, "visuals")
MODELS_DIR  = os.path.join(RESULTS_DIR, "models")
METRICS_DIR = os.path.join(RESULTS_DIR, "metrics")

os.makedirs(VISUALS_DIR, exist_ok=True)
os.makedirs(MODELS_DIR,  exist_ok=True)
os.makedirs(METRICS_DIR, exist_ok=True)


# =========================================================
# LOAD SPLITS
# =========================================================

X_train = np.load(os.path.join(SPLIT_DIR, "X_train.npy"))
y_train = np.load(os.path.join(SPLIT_DIR, "y_train.npy"))

X_val   = np.load(os.path.join(SPLIT_DIR, "X_val.npy"))
y_val   = np.load(os.path.join(SPLIT_DIR, "y_val.npy"))

X_test  = np.load(os.path.join(SPLIT_DIR, "X_test.npy"))
y_test  = np.load(os.path.join(SPLIT_DIR, "y_test.npy"))

# ---------------------------------------------------------
# Merge train + val for final model training.
#
# CV runs on X_tv (train+val) with internal folds,
# and the final model also trains on X_tv — both see the
# same data budget, so CV and test accuracy are comparable.
# ---------------------------------------------------------

X_tv = np.concatenate([X_train, X_val], axis=0)
y_tv = np.concatenate([y_train, y_val], axis=0)

print(f"✓ Data loaded | Train+Val: {X_tv.shape} | Test: {X_test.shape}")


# =========================================================
# CLASS NAMES
# =========================================================

class_names = ["Demented", "Nondemented", "Converted"]


# =========================================================
# MODELS
# ---------------------------------------------------------
# FIX 2: Wrap every model in a Pipeline(scaler + model).
#
# Before: scaler was fit once on X_train and applied to all
# splits. Inside cross_val_score the model sees raw features
# from different folds, but the scaler statistics (mean/std)
# came from the full X_train — that is data leakage.
#
# With Pipeline, sklearn re-fits the scaler on each CV
# training fold and transforms the val fold using ONLY
# those fold's statistics. No leakage possible.
#
# FIX 3: Add mild regularization to tree-based models.
# Default DecisionTree has unlimited depth — it memorizes
# all 260 training rows perfectly (100% train acc) but
# generalizes poorly. max_depth=8 still captures the
# signal without memorizing noise.
# =========================================================

def make_pipeline(estimator):
    """Wrap estimator in StandardScaler pipeline."""
    return Pipeline([
        ("scaler", StandardScaler()),
        ("model",  estimator),
    ])


pipelines = {

    "Logistic Regression": make_pipeline(
        LogisticRegression(
            C=1.0,           # L2 penalty, tunable
            max_iter=1000,
            random_state=42,
            class_weight="balanced",   # handles Converted imbalance
        )
    ),

    "Decision Tree": make_pipeline(
        DecisionTreeClassifier(
            max_depth=8,             # FIX: was unlimited → memorized data
            min_samples_leaf=3,      # FIX: at least 3 samples per leaf
            class_weight="balanced",
            random_state=42,
        )
    ),

    "Random Forest": make_pipeline(
        RandomForestClassifier(
            n_estimators=200,
            max_depth=10,
            min_samples_leaf=2,
            class_weight="balanced",
            random_state=42,
            n_jobs=-1,
        )
    ),

    "Gradient Boosting": make_pipeline(
        GradientBoostingClassifier(
            n_estimators=150,
            learning_rate=0.05,       # lower LR → better generalization
            max_depth=4,
            subsample=0.8,            # row subsampling reduces overfitting
            random_state=42,
        )
    ),

    "XGBoost": make_pipeline(
        XGBClassifier(
            n_estimators=150,
            learning_rate=0.05,
            max_depth=4,
            subsample=0.8,
            colsample_bytree=0.8,
            reg_alpha=0.1,            # L1 regularization
            reg_lambda=1.0,           # L2 regularization
            eval_metric='mlogloss',
            random_state=42,
            verbosity=0,
            use_label_encoder=False,
        )
    ),

    "SVM": make_pipeline(
        SVC(
            C=1.0,
            kernel="rbf",
            probability=True,
            class_weight="balanced",
            random_state=42,
        )
    ),

    "KNN": make_pipeline(
        KNeighborsClassifier(
            n_neighbors=7,            # slightly more neighbors = smoother boundary
            weights="distance",
        )
    ),

    "Naive Bayes": make_pipeline(
        GaussianNB()
    ),
}


# =========================================================
# CROSS-VALIDATION + FINAL TRAINING
# ---------------------------------------------------------
# FIX 4: CV runs on X_tv (train+val combined).
#
# Before: CV ran only on X_train (70% of data). The final
# model then trained on the same X_train. So CV accuracy
# was computed on a 70% subset but reported as if it
# represented the full generalization — misleading.
#
# Now both steps use X_tv consistently.
# =========================================================

cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)

results      = {}
conf_matrices = {}

for name, pipeline in pipelines.items():

    print(f"\nTraining: {name}")

    t0 = time.time()

    # --- cross-validate on train+val (no leakage — Pipeline handles scaling) ---
    cv_output = cross_validate(
        pipeline,
        X_tv,
        y_tv,
        cv=cv,
        scoring="accuracy",
        return_train_score=True,   # expose train acc to spot remaining overfit
        n_jobs=-1,
    )

    # --- final model: fit on full train+val ---
    pipeline.fit(X_tv, y_tv)

    train_time = time.time() - t0

    # --- evaluate on held-out test set ---
    y_pred = pipeline.predict(X_test)
    y_prob = pipeline.predict_proba(X_test)

    acc = accuracy_score(y_test, y_pred)
    f1  = f1_score(y_test, y_pred, average="weighted")
    auc = roc_auc_score(y_test, y_prob, multi_class="ovr", average="weighted")

    cv_train_mean = cv_output["train_score"].mean()
    cv_val_mean   = cv_output["test_score"].mean()
    cv_val_std    = cv_output["test_score"].std()

    results[name] = {
        "Accuracy":        round(acc,            4),
        "F1 (weighted)":   round(f1,             4),
        "AUC-ROC (OvR)":   round(auc,            4),
        "CV Acc (mean)":   round(cv_val_mean,    4),
        "CV Acc (std)":    round(cv_val_std,     4),
        "CV Train Acc":    round(cv_train_mean,  4),  # for overfit diagnosis
        "Train Time (s)":  round(train_time,     3),
    }

    conf_matrices[name] = confusion_matrix(y_test, y_pred)

    # save pipeline (includes scaler — ready for inference)
    model_path = os.path.join(MODELS_DIR, f"{name.replace(' ', '_')}_pipeline.pkl")
    joblib.dump(pipeline, model_path)

    overfit_gap = cv_train_mean - cv_val_mean
    flag = " ⚠ overfit" if overfit_gap > 0.10 else ""

    print(
        f"  ✓ {name:22s} | "
        f"Test={acc:.3f} | CV={cv_val_mean:.3f}±{cv_val_std:.3f} | "
        f"Gap={overfit_gap:+.3f}{flag}"
    )


# =========================================================
# RESULTS DATAFRAME
# =========================================================

results_df = (
    pd.DataFrame(results)
    .T
    .reset_index()
    .rename(columns={"index": "Model"})
    .sort_values("Accuracy", ascending=False)
    .reset_index(drop=True)
)

print("\n" + "=" * 70)
print("FINAL RESULTS (sorted by accuracy)")
print("=" * 70)
print(results_df[["Model", "Accuracy", "F1 (weighted)", "AUC-ROC (OvR)",
                   "CV Acc (mean)", "CV Acc (std)", "Train Time (s)"]].to_string(index=False))


# =========================================================
# FIGURE 3 — MODEL PERFORMANCE DASHBOARD
# =========================================================

model_names  = results_df["Model"].tolist()
colors_bar   = plt.cm.tab10(np.linspace(0, 1, len(model_names)))
x_pos        = np.arange(len(model_names))

fig3 = plt.figure(figsize=(22, 20))
fig3.suptitle("Classical Model Performance", fontsize=18, fontweight="bold")
gs3 = gridspec.GridSpec(3, 3, figure=fig3, hspace=0.6, wspace=0.4)

# --- Test accuracy ---
ax = fig3.add_subplot(gs3[0, 0])
ax.barh(model_names, results_df["Accuracy"].values, color=colors_bar)
ax.set_xlim(0, 1.05)
ax.set_title("Test Accuracy")

# --- F1 score ---
ax = fig3.add_subplot(gs3[0, 1])
ax.barh(model_names, results_df["F1 (weighted)"].values, color=colors_bar)
ax.set_xlim(0, 1.05)
ax.set_title("Weighted F1 Score")

# --- AUC ---
ax = fig3.add_subplot(gs3[0, 2])
ax.barh(model_names, results_df["AUC-ROC (OvR)"].values, color=colors_bar)
ax.set_xlim(0, 1.05)
ax.set_title("AUC-ROC")

# --- CV accuracy with train/val bars side by side (overfit view) ---
ax = fig3.add_subplot(gs3[1, 0])
cv_vals   = results_df["CV Acc (mean)"].values
cv_trains = results_df["CV Train Acc"].values
cv_stds   = results_df["CV Acc (std)"].values
w = 0.35
ax.bar(x_pos - w/2, cv_trains, w, label="CV Train",  color="#1565C0", alpha=0.7)
ax.bar(x_pos + w/2, cv_vals,   w, label="CV Val",    color=colors_bar, yerr=cv_stds, capsize=4)
ax.set_xticks(x_pos)
ax.set_xticklabels([m.replace(" ", "\n") for m in model_names], fontsize=8)
ax.set_ylim(0, 1.15)
ax.legend(fontsize=8)
ax.set_title("CV Train vs Val Accuracy (gap = overfit)")

# --- Training time ---
ax = fig3.add_subplot(gs3[1, 1])
ax.bar(x_pos, results_df["Train Time (s)"].values, color=colors_bar)
ax.set_xticks(x_pos)
ax.set_xticklabels([m.replace(" ", "\n") for m in model_names], fontsize=8)
ax.set_title("Training Time (s)")

# --- Top 3 comparison ---
top3 = results_df.head(3)["Model"].tolist()
metrics_compare = ["Accuracy", "F1 (weighted)", "AUC-ROC (OvR)", "CV Acc (mean)"]
ax = fig3.add_subplot(gs3[1, 2])
x = np.arange(len(metrics_compare))
w = 0.25
for i, m in enumerate(top3):
    vals = [results[m][k] for k in metrics_compare]
    ax.bar(x + i * w, vals, w, label=m)
ax.set_xticks(x + w)
ax.set_xticklabels(metrics_compare, rotation=15, fontsize=8)
ax.set_ylim(0, 1.15)
ax.legend(fontsize=8)
ax.set_title("Top 3 Models")

# --- Confusion matrices ---
for i, m_name in enumerate(top3):
    ax = fig3.add_subplot(gs3[2, i])
    sns.heatmap(
        conf_matrices[m_name],
        annot=True, fmt="d", cmap="Blues",
        xticklabels=class_names,
        yticklabels=class_names,
        cbar=False, ax=ax,
    )
    ax.set_title(m_name)
    ax.set_xlabel("Predicted")
    ax.set_ylabel("Actual")

fig3_path = os.path.join(VISUALS_DIR, "fig3_model_results.png")
plt.savefig(fig3_path, dpi=300, bbox_inches="tight")
plt.close()
print(f"\n✓ Figure 3 saved → {fig3_path}")


# =========================================================
# FIGURE 4 — FEATURE IMPORTANCE
# =========================================================

rf_model = pipelines["Random Forest"].named_steps["model"]

feature_labels = ["Age", "Sex", "EDUC", "SES", "MMSE",
                  "CDR", "eTIV", "nWBV", "ASF", "MR Delay"]

fi = (
    pd.Series(rf_model.feature_importances_, index=feature_labels)
    .sort_values()
)

fig4, ax = plt.subplots(figsize=(10, 6))
ax.barh(fi.index, fi.values, color="#1565C0")
ax.set_title("Random Forest Feature Importance")
ax.set_xlabel("Importance")

fig4_path = os.path.join(VISUALS_DIR, "fig4_feature_importance.png")
plt.savefig(fig4_path, dpi=300, bbox_inches="tight")
plt.close()
print(f"✓ Figure 4 saved → {fig4_path}")


# =========================================================
# SAVE EXCEL RESULTS
# =========================================================

excel_path = os.path.join(METRICS_DIR, "model_results.xlsx")

wb = Workbook()
ws = wb.active
ws.title = "Model Results"

# Exclude internal diagnostic column from the exported table
export_cols = ["Model", "Accuracy", "F1 (weighted)", "AUC-ROC (OvR)",
               "CV Acc (mean)", "CV Acc (std)", "Train Time (s)"]

export_df = results_df[export_cols]
headers   = export_cols

for col_i, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_i, value=h)
    cell.font      = Font(bold=True, color="FFFFFF")
    cell.fill      = PatternFill("solid", start_color="1565C0")
    cell.alignment = Alignment(horizontal="center")

for row in export_df.itertuples(index=False):
    ws.append(list(row))

# Auto-width columns
for col in ws.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    ws.column_dimensions[col[0].column_letter].width = max_len + 4

wb.save(excel_path)
print(f"✓ Excel saved → {excel_path}")

print("\n=== ALL DONE ===")